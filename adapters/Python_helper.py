from bs4 import BeautifulSoup
import re, json
import html
import os
from datetime import datetime
from robot.api import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tempfile
import win32com.client
from PyPDF2 import PdfMerger
import pygetwindow as gw
import pyautogui
import time
from robot.api import logger
from pathlib import Path
import os
import shutil
import ctypes
from ctypes import wintypes
from contextlib import contextmanager
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.edge.options import Options
import time



# -----------------------------
# Win32 (ctypes) setup
# -----------------------------
FORMAT_MESSAGE_FROM_SYSTEM = 0x00001000
LOGON32_LOGON_INTERACTIVE = 2          # Suitable for filesystem access
LOGON32_PROVIDER_DEFAULT = 0

advapi = ctypes.WinDLL("advapi32", use_last_error=True)
kernel = ctypes.WinDLL("kernel32", use_last_error=True)

advapi.LogonUserW.argtypes = [
    wintypes.LPCWSTR,  # lpszUsername
    wintypes.LPCWSTR,  # lpszDomain
    wintypes.LPCWSTR,  # lpszPassword
    wintypes.DWORD,    # dwLogonType
    wintypes.DWORD,    # dwLogonProvider
    ctypes.POINTER(wintypes.HANDLE)  # phToken
]
advapi.LogonUserW.restype = wintypes.BOOL

advapi.ImpersonateLoggedOnUser.argtypes = [wintypes.HANDLE]
advapi.ImpersonateLoggedOnUser.restype = wintypes.BOOL

advapi.RevertToSelf.argtypes = []
advapi.RevertToSelf.restype = wintypes.BOOL

kernel.CloseHandle.argtypes = [wintypes.HANDLE]
kernel.CloseHandle.restype = wintypes.BOOL

kernel.FormatMessageW.argtypes = [
    wintypes.DWORD, wintypes.LPCVOID, wintypes.DWORD,
    wintypes.DWORD, wintypes.LPWSTR, wintypes.DWORD, wintypes.LPVOID
]
kernel.FormatMessageW.restype = wintypes.DWORD

def open_url(url):
    options = Options()
    options.add_argument("--start-maximized")

    driver = webdriver.Edge(options=options)
    driver.get(url)
    time.sleep(2)
    driver.close()
    return driver

def extract_header_details(html_content: str):
    if "<table" in html_content:
        header_html = html_content.split("<table", 1)[0]
    else:
        header_html = html_content
    decoded     = html.unescape(header_html)
    soup        = BeautifulSoup(decoded, "html.parser")
    header_text = soup.get_text("\n").replace("\xa0", " ")
    lines       = [l.strip() for l in header_text.split("\n") if l.strip()]

    remitter_name = None
    for line in lines:
        if "payment settlement list" in line.lower():
            remitter_name = line.lower().split("payment settlement list")[0].strip()
            break

    value_date = None
    for line in lines:
        m = re.search(r"(\d{2}\.\d{2}\.\d{4})\s*/\s*\d{2}:\d{2}:\d{2}", line)
        if m:
            value_date = m.group(1)
            break

    company_code = None
    for line in lines:
        m = re.search(r"Company Code:\s*(\d+)", line)
        if m:
            company_code = m.group(1)
            break

    remitter_address = None
    for line in lines:
        if "Users:" in line:
            remitter_address = line.split("Users:")[0].strip()
            break

    return {
        "value_date":       value_date,
        "company_code":     company_code,
        "remitter_name":    remitter_name,
        "remitter_address": remitter_address,
    }

def extract_account_details(html_content: str):
    soup = BeautifulSoup(html_content, "html.parser")
    for table in soup.find_all("table"):
        table_text = table.get_text(" ", strip=True).replace("\xa0", " ")
        if "Our account number" in table_text:
            rows = table.find_all("tr")
            if len(rows) < 2:
                return None, None
            header_cells = [td.get_text(strip=True).replace("\xa0", " ") for td in rows[0].find_all("td")]
            data_cells   = [td.get_text(strip=True).replace("\xa0", " ") for td in rows[1].find_all("td")]
            try:
                acc_index  = header_cells.index("Our account number")
                curr_index = header_cells.index("Crcy")
            except ValueError:
                return None, None
            return data_cells[acc_index], data_cells[curr_index]
    return None, None

def extract_vendors(soup):
    vendors, cur = [], None

    for el in soup.find_all(["span", "table"]):
        if el.name == "span" and "white-space:nowrap" in (el.get("style") or ""):
            t = re.sub(r"\s+", " ", "".join(n.get_text() for n in el.find_all("nobr")).strip())

            m = re.search(r"--(Vendor|Supplier)\s+(\d+)", t, re.I)
            if m:
                if cur: vendors.append(_finish(cur))
                cur = {
                    "supplier_number":     m.group(2),
                    "beneficiary_name":    None,
                    "beneficiary_address": None,
                    "swift_name":          None,
                    "swift_number":        None,
                    "iban":                None,
                    "account_number":      None,
                    "bank_number":         None,
                    "KZ":                  None,
                    "currency":            None,
                    "amount":              None,
                    "barcodes":            [],
                    "sub_rows":            [],
                    "_vendor_lines":       [],
                    "_bank_lines":         [],
                }
                continue

            if not cur: continue

            if "|" in t:
                parts = t.split("|")
                left  = parts[1].strip() if len(parts) > 1 else ""
                right = parts[3].strip() if len(parts) > 3 else ""
                if left:  cur["_vendor_lines"].append(left.rstrip(">").strip())
                if right: cur["_bank_lines"].append(right.rstrip(">").strip())

        elif el.name == "table" and cur:
            for tr in el.find_all("tr"):
                t = ""
                for font in tr.find_all("font"):
                    if "courier" in (font.get("face") or "").lower():
                        t += " ".join(n.get_text() for n in font.find_all("nobr"))
                t = re.sub(r"\s+", " ", t).strip()
                if not t: continue

                if cur["KZ"] is None and re.search(r"\d{6,12}\s+[A-Z]{2,4}\s+[A-Z]{3}", t):
                    m = re.search(r"(\d{6,12})\s+[A-Z]{2,4}\s+[A-Z]{3}", t)
                    cur["KZ"] = m.group(1) if m else None
                    m = re.search(r"\b([A-Z]{3})\s*$", t)
                    cur["currency"] = m.group(1) if m else None
                    m = re.search(r"([\d.,]+-?)\s+[A-Z]{3}\s*$", t)
                    if m: cur["amount"] = m.group(1)
                    continue

                if re.match(r"^\*", t): continue
                m = re.search(r"(\d{4})\s+(\d+)\s+([A-Z]{2})\s+(\d{2}\.\d{2}\.\d{4})", t)
                if m:
                    barcode = m.group(2)
                    cur["sub_rows"].append(barcode)
                    cur["barcodes"].append(barcode)

    if cur: vendors.append(_finish(cur))
    return vendors

def _finish(cur):
    vlines = cur.pop("_vendor_lines", [])
    blines = cur.pop("_bank_lines", [])

    cur["beneficiary_name"] = vlines[0] if vlines else None
    addr = " ".join(vlines[1:])
    addr = re.sub(r",+", ",", addr)
    addr = re.sub(r" +", " ", addr).strip()
    cur["beneficiary_address"] = addr if addr else None

    cur["swift_name"] = blines[0] if blines else None

    for line in blines:
        if not cur["swift_number"] and re.search(r"SWIFT|BIC", line, re.I):
            m = re.search(r"(?:SWIFT|BIC)[:\s]*(\S+)", line, re.I)
            if m: cur["swift_number"] = m.group(1)
        if not cur["iban"] and re.search(r"IBAN", line, re.I):
            m = re.search(r"IBAN[:\s]+(.+)", line, re.I)
            if m: cur["iban"] = m.group(1).strip()
        if not cur["bank_number"] and re.search(r"BNo|Bank.?No|Bank Number", line, re.I):
            m = re.search(r"(?:BNo|Bank[\s.]?No[\w.]*|Bank Number)[:\s]+(\S+)", line, re.I)
            if m: cur["bank_number"] = m.group(1)
        if not cur["account_number"] and re.search(r"Acct|Account", line, re.I):
            m = re.search(r"(?:Acct|Account[\s\w.]*)[:\s]+(\S+)", line, re.I)
            if m: cur["account_number"] = m.group(1)

    return cur

def extract_all(filepath: str) -> dict:
 
    
    html_content = open(filepath, encoding="utf-8", errors="ignore").read()
    soup         = BeautifulSoup(html_content, "html.parser")

    header               = extract_header_details(html_content)
    our_account, our_currency = extract_account_details(html_content)
    vendors              = extract_vendors(soup)

    return {
        "value_date":       header["value_date"],
        "company_code":     header["company_code"],
        "remitter_name":    header["remitter_name"],
        "remitter_address": header["remitter_address"],
        "remmiter_account":      our_account,
        "remmiter_account_currency":     our_currency,
        "vendors":          vendors,
    }

def create_folder(path):
    os.makedirs(path, exist_ok=True)
    print(f"  Created: {path}")

def create_cover_page(folder_path, vendor, header):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45

    thin = Side(style="thin")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # [row, merge_a, col_a_label, col_b, col_c_value]
    rows = [
        [1,  "A1:A2", "Amount to be remitted (in Foreign Currency)", "Currency",              vendor["currency"]],
        [2,  None,    None,                                           "Amount",                vendor["amount"]],
        [3,  "A3:A4", "Remitter's name and address",                  "Name",                  header["remitter_name"]],
        [4,  None,    None,                                           "Address",               header["remitter_address"]],
        [5,  None,    "Remitter's bank account no",                   "Account no of buyer",   header["remmiter_account"]],
        [6,  "A6:A7", "Beneficiary's name and address",               "Name",                  vendor["beneficiary_name"]],
        [7,  None,    None,                                           "Address",               vendor["beneficiary_address"]],
        [8,  None,    "Beneficiary's bank account no",                "Account no of supplier",vendor["iban"] or vendor["account_number"]],
        [9,  "A9:A10","Beneficiary's bank name and SWIFT",            "Name",                  vendor["swift_name"]],
        [10, None,    None,                                           "SWIFT code",            vendor["swift_number"]],
        [11, None,    "Payment run date",                             "Value date",            header["value_date"]],
        [12, None,    "Transfer reference number",                    "End to end ID",         vendor["KZ"]],
    ]

    for r, merge, label, b_val, c_val in rows:
        if merge: ws.merge_cells(merge)
        for col, val in [(1, label), (2, b_val), (3, c_val)]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = brd
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.font = Font(name="Arial", size=10, bold=(col == 1 and val))

    filepath = os.path.join(folder_path, "CoverPageTemplate.xlsx")
    wb.save(filepath)
    print(f"  Excel saved: {filepath}")
    return filepath

def extract_data(File , PATH_ARCHIVES , PREFIX):
    # ── extract everything ────────────────────────────────────────
    data = extract_all(File)

    company_code      = data["company_code"]
    remitter_name     = data["remitter_name"]
    remitter_address  = data["remitter_address"]
    value_date        = data["value_date"]
    remmiter_account  = data["remmiter_account"]
    remmiter_currency = data["remmiter_account_currency"]
    vendors           = data["vendors"]

    print(f"\nCompany Code       : {company_code}")
    print(f"Remitter Name      : {remitter_name}")
    print(f"Remitter Address   : {remitter_address}")
    print(f"Value Date         : {value_date}")
    print(f"Remitter Account   : {remmiter_account}")
    print(f"Remitter Currency  : {remmiter_currency}")
    print(f"Vendors Found      : {len(vendors)}\n")

    # ── Step 28: base folder  PATH_ARCHIVES\01.1_VHVA 4211 ───────
    company_folder = os.path.join(PATH_ARCHIVES, f"{PREFIX} {company_code}")
    create_folder(company_folder)
    logger.console(data)
    barcodes = data.get("barcodes", [])


    for vendor in vendors:
        beneficiary_name = vendor["beneficiary_name"] or "UNKNOWN"
        currency         = vendor["currency"]         or ""
        amount           = vendor["amount"]           or ""
        kz               = vendor["KZ"]               or ""
        date_clean       = value_date.replace(".", "") if value_date else ""

        # ── Step 28: subfolder per beneficiary ───────────────────
        beneficiary_folder = os.path.join(company_folder, beneficiary_name)
        create_folder(beneficiary_folder)
        barcodes = vendor.get("barcodes", [])

        # ── Step 29: beneficiary + currency + amount ──────────────
        currency_amount_folder = os.path.join(
            beneficiary_folder,
            f"{beneficiary_name} {currency}-{amount}"
        )
        create_folder(currency_amount_folder)

        # ── Step 30: company + KZ + date + beneficiary ────────────
        kz_folder = os.path.join(
            currency_amount_folder,
            f"{company_code}{kz}{date_clean}_{beneficiary_name}"
        )
        create_folder(kz_folder)

        # step 31 
        cover_page_path = create_cover_page(PATH_ARCHIVES , vendor, data)
        return  kz_folder , cover_page_path

def close_edge_pdf_windows():
    # Get all windows
    windows = gw.getWindowsWithTitle('')
    time.sleep(2)
    for w in windows:
        title = w.title
        # Check if it's an Edge window containing .pdf
        if 'Edge' in title and '.pdf' in title:
            print(f'Closing window: {title}')
            w.activate()          # Bring to foreground
            time.sleep(0.5)
            pyautogui.hotkey('alt', 'f4')

def monitor_windows(pdf_name_part, duration_seconds=10, interval_seconds=1, close=True):

    start = time.time()
    logger.console(pdf_name_part)
    logger.console(f"\nMonitoring all windows for {duration_seconds} seconds...\n")
    
    while time.time() - start < duration_seconds:
        windows = gw.getAllWindows()
        logger.console(f"\nTime: {round(time.time() - start, 1)}s - {len(windows)} windows found")
        
        for w in windows:
            title_lower = w.title.lower()
            logger.console(f"  Window title: '{w.title}'")
            # logger.console(pdf_name_part[:10] , pdf_name_part)
            # Check if PDF name part is in window title

            if pdf_name_part.lower() in title_lower:
                logger.console(f"    >>> MATCH FOUND for '{pdf_name_part}' in '{w.title}'")
                
                if close:
                    try:
                        w.activate()             # Bring window to foreground
                        time.sleep(0.5)  
                        os.system(f'taskkill /F /FI "WINDOWTITLE eq {w.title}*" /T')
                        logger.console(f"  >>> Closed window: '{w.title}'")
                    except Exception as e:
                        logger.console(f"    >>> Failed to close window: {e}")
        
        time.sleep(interval_seconds)

def wait_until_file_free(filepath, timeout=30):
    start = time.time()
    logger.console("HEre in Wait for file free status")
    while True:
        try:
            logger.console("file is free now")
            os.rename(filepath, filepath)  # test if file is locked
            return True
        except PermissionError:
            logger.console("HEre in Wait until file fro free")
            if time.time() - start > timeout:
                raise Exception(f"File still locked: {filepath}")
            time.sleep(2)

def merge_cover_excel_with_kz_data(data, cover_excel_path, kz_folder, path_archives):
    """
    Converts Excel cover to PDF, merges it with all barcode PDFs in kz_folder,
    and saves the merged PDF at dynamic path using data from ${data}.
    """
    vendor = data['vendors'][0]  # assuming single vendor per KZ

    company_code = data['company_code']
    beneficiary_name = vendor['beneficiary_name']
    currency = vendor['currency']
    amount = vendor['amount']
    kz_number = vendor['KZ']
    value_date = data['value_date']
    transaction_ref = vendor['KZ']  # or use some regex transaction if different

    # Build dynamic folder path
    folder_path = os.path.join(
        path_archives,
        company_code,
        beneficiary_name,
        f"{beneficiary_name} {currency}-{amount}",
        f"{company_code}{kz_number}{value_date}{beneficiary_name}"
    )
    os.makedirs(folder_path, exist_ok=True)

    # Final merged PDF path
    merged_pdf_path = os.path.join(folder_path, f"{transaction_ref}_{beneficiary_name}.pdf")

    # Collect all PDF files in KZ folder
    pdf_files = [os.path.join(kz_folder, f) for f in os.listdir(kz_folder) if f.lower().endswith('.pdf')]
    if not pdf_files:
        raise ValueError(f"No PDF files found in {kz_folder}")
    pdf_files.sort(key=os.path.getmtime)

    # Convert Excel cover to temp PDF
    temp_dir = tempfile.mkdtemp()
    cover_pdf_temp = os.path.join(temp_dir, "cover_temp.pdf")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(cover_excel_path)
    wb.ExportAsFixedFormat(0, cover_pdf_temp)
    wb.Close(False)
    excel.Quit()

    # Merge cover + barcode PDFs
    merger = PdfMerger()
    merger.append(cover_pdf_temp)
    for pdf in pdf_files:
        merger.append(pdf)
    merger.write(merged_pdf_path)
    merger.close()

    # Clean up temporary cover
    os.remove(cover_pdf_temp)
    os.rmdir(temp_dir)

    return merged_pdf_path

def extract_robot_data(filepath, barcode_folder):

    html_content = open(filepath, encoding="utf-8", errors="ignore").read()

    # safety check — only process if "4 CGI payments" is present
    is_cgi_file = bool(re.search(r"4(?:\s|&nbsp;)+CGI(?:\s|&nbsp;)+payments(?:\s|&nbsp;)+abroad", html_content, re.I))



    if not is_cgi_file:
        print(f"SKIPPED: '4 CGI payments' not found in {filepath}")
        return None, [], None, None, False
    
    data = extract_all(filepath)

    value_date = data["value_date"]
    remitter_account = data["remmiter_account"]
    vendors = data["vendors"]

    # collect all barcodes
    # collect all barcodes
    kz_barcode_list = []
    for vendor in vendors:
        barcodes = vendor.get("barcodes", [])
        kz = vendor.get("KZ")

        for b in barcodes:
            kz_barcode_list.append({
                "KZ": kz,
                "barcode": b
            })
    
    if barcode_folder:
        os.makedirs(barcode_folder, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(barcode_folder, f"BarcodeList.txt")
        
        with open(output_file, "w", encoding="utf-8") as f:
            for item in kz_barcode_list:
                barcode = item.get("barcode")
                if barcode:
                    f.write(f"{barcode}\n")
        
        logger.console(f"Barcodes saved to: {output_file}")

    return value_date, kz_barcode_list, remitter_account, data, True

def _format_win_error(err=None) -> str:
    """Return a Windows error as human-readable text."""
    if err is None:
        err = ctypes.get_last_error()
    buf = ctypes.create_unicode_buffer(2048)
    kernel.FormatMessageW(
        FORMAT_MESSAGE_FROM_SYSTEM,
        None,
        err,
        0,
        buf,
        len(buf),
        None
    )
    msg = buf.value.strip()
    return f"[{err}] {msg}" if msg else f"[{err}] Unknown error"


@contextmanager
def impersonate(username: str, password: str, domain: str | None = None):
    """
    Temporarily impersonate a Windows user using LogonUserW/ImpersonateLoggedOnUser.
    - domain: AD domain name, computer name, or '.' for local accounts.
    """
    if os.name != "nt":
        raise OSError("This function only works on Windows.")
    if not domain:
        domain = "."

    token = wintypes.HANDLE()
    ok = advapi.LogonUserW(
        username, domain, password,
        LOGON32_LOGON_INTERACTIVE,
        LOGON32_PROVIDER_DEFAULT,
        ctypes.byref(token)
    )
    print(username , domain , password)
    if not ok:
        raise OSError(f"LogonUserW failed: {_format_win_error()}")

    try:
        if not advapi.ImpersonateLoggedOnUser(token):
            raise OSError(f"ImpersonateLoggedOnUser failed: {_format_win_error()}")
        try:
            yield
        finally:
            if not advapi.RevertToSelf():
                raise OSError(f"RevertToSelf failed: {_format_win_error()}")
    finally:
        if token:
            kernel.CloseHandle(token)


def move_from_myuser_download_to_temp(
    cred_user: str,
    cred_password: str,
    cred_domain: str | None = None,
    source_dir: str = r"C:\Users\z12315\Downloads",
    target_dir: str = r"C:\TEMP\RPA0024-VTI",
    overwrite: bool = False,
    delete_source: bool = True
) -> str:
    if os.name != "nt":
        raise OSError("Windows only.")

    source_dir = os.path.normpath(source_dir)
    target_dir = os.path.normpath(target_dir)

    # Step 1: Perform operations under impersonation
    with impersonate(cred_user, cred_password, cred_domain):
        # Ensure target exists
        os.makedirs(target_dir, exist_ok=True)

        if not os.path.exists(source_dir):
            print(f"Source directory {source_dir} not found!")
            return ""

        # Iterate through all items in source
        for filename in os.listdir(source_dir):
            src_path = os.path.join(source_dir, filename)
            dst_path = os.path.join(target_dir, filename)

            if os.path.isfile(src_path):
                if os.path.exists(dst_path):
                    if not overwrite:
                        continue 
                    os.remove(dst_path)

                shutil.copy2(src_path, dst_path)
                if delete_source:
                    os.remove(src_path)
                print(f"Successfully moved: {filename}")

    return target_dir

  